# -*- coding:utf-8 -*-
import os
import copy
import re

import shutil
import sys
import zipfile

from flask import current_app, session
from openpyxl import Workbook, load_workbook
from robot.api import TestSuiteBuilder

from utils.file import remove_dir

from utils.mylogger import getlogger

log = getlogger(__name__)

def getCaseDoc_pytest(cpath, cname):
    """
    通过模块导入，得到 pytest 用的 doc ，至于代码，需要后续引入
    :param cpath:
    :param cname:
    :return:
    """
    case_dir = os.path.dirname(cpath)
    case_file = os.path.basename(cpath)
    case_module = os.path.splitext(case_file)[0]
    case_name = cname

    if case_name.find("::") != -1:
        case_class = case_name.split("::")[0]
        case_name = case_name.split("::")[1]
    else:
        case_class = None

    sys.path.insert(0, case_dir)
    cm = __import__(case_module)

    content = "Not found"

    if not case_class:
        fun = cm.__dict__.get(case_name, None)
        if not fun:
            log.info("找不到函数:{}".format(case_name))
            return "Not found"
        return fun.__doc__

    else:
        cls = cm.__dict__.get(case_class, None)
        if not cls:
            log.info("找不到类:{}".format(case_class))
            return "Not found"
        fun = cls.__dict__.get(case_name)
        if not fun:
            log.info("类{}中找不到函数:{}".format(case_class, case_name))
            return "Not found"
        return fun.__doc__

    sys.path.pop()
    return "Not found"

def getCaseContent(cpath, cname):
    """
    直接通过文本格式进行字符串的处理， 基于RF的用例文件格式
    :param cpath: 文件名
    :param cname: 用例名
    :return: 用例内容
    """
    if not os.path.exists(cpath):
        return "Can not find case file:"+cpath

    content = ''
    case_name = cname.strip()
    case_start = False
    content_start = False
    with open(cpath, 'r') as cf:
        for line in cf.readlines():
            if line.startswith("*** Test"):
                case_start = True
                continue
            if case_start and line.strip() == case_name:
                content_start = True
                continue
            if case_start and content_start and line.startswith(' '*4):
                content += line.strip() + "\r\n"
                continue
            if case_start and content_start and line.strip() == "":
                content += line.strip() + "\r\n"
                continue
            if case_start and content_start and (not line.startswith(' '*4)):
                return content

def export_casezip(key, exp_filedir=''):

    dir = exp_filedir
    if dir == '':
        dir = os.environ["PROJECT_DIR"] + '/runtime'

    zip_name = os.path.basename(key) + '.zip'
    zip_path = os.path.join(dir, zip_name)

    try:
        z = zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED)

        for dirpath, dirnames, filenames in os.walk(key):
            fpath = dirpath.replace(key, '')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                z.write(os.path.join(dirpath, filename), fpath + filename)
        z.close()
    except Exception as e:
        log.error("下载zip用例异常:{}".format(e))
        return (False, "{}".format(e))

    return (True, zip_path)


def export_casexlsx(key, db, exp_filedir=''):
    """
    Download Readable test case
    :param key: Dir of the case
    :param db: case info db
    :param exp_filedir: output file dir
    :return: result, file
    """

    export_dir = key
    if not os.path.isdir(export_dir):
        log.error("不支持导出一个文件中的用例:"+export_dir)
        return (False, "不支持导出一个文件中的用例:"+export_dir)

    basename = os.path.basename(export_dir)

    dir = exp_filedir
    if dir == '':
        dir = os.environ["AUTO_TEMP"]

    os.mkdir(dir) if not os.path.exists(dir) else None

    export_file = os.path.join(dir, basename+'.xlsx')

    db.refresh_caseinfo(export_dir, "Force")

    cases = []
    sql = "SELECT info_key,info_name,info_doc,info_tags FROM testcase WHERE info_key like '{}%' ;".format(
        key)
    res = db.runsql(sql)
    for i in res:
        (info_key, info_name, info_doc, info_tag) = i
        cases.append([info_key, info_name, info_doc, info_tag])

    wb = Workbook()

    ws = wb.active
    ws.append(["导出&导入用例："])
    ws.append(["'_'用在文件名前后，表示用例文件：如 '_用例文件名_' 表示'用例文件名.robot'"])
    ws.append(["'-'用来连接目录，没有此符号表示没有子目录：如 '目录1-目录11' 表示 '目录1/目录11'"])
    ws.append(["每个sheet的第一列，是后面用例所在的用例文件名（.robot）"])
    ws.append(["... ..."])
    ws.append(["注意：通过xlsx文件导入用例，如果是自动化用例，且用例已经存在，则只更新doc和tag，不更新用例内容"])
    ws.append(["... ..."])
    ws.append(["此'sheet'页面，不会被导入"])
    ws.append(["... ..."])
    ws.append(["Export&Import Cases："])
    ws.append(
        ["'_'after the file name，Means a suite：'_SuiteName_' means 'SuiteName.robot'"])
    ws.append(
        ["'-'concat the dirs，no this sign no subdir：'dir1-dir11' means 'dir1/dir11'"])
    ws.append(
        ["First Column of each sheet，is the suite name of the case in this line（.robot）"])
    ws.append(["... ..."])
    ws.append(["Caution：Import cases from xlsx file，if it is Auto-case and it exists，then update doc and tag Only，Do not update Case content."])
    ws.append(["... ..."])
    ws.append(["This 'sheet' ，Wont be imported."])

    for c in cases:
        if not os.path.exists(c[0]):
            continue

        casecontent = getCaseContent(c[0], c[1])
        suitename = os.path.basename(c[0])

        tags = c[3].split(',')
        tags.remove('${EMPTY}') if '${EMPTY}' in tags else None

        category = "Auto"

        if "HAND" in tags or "Hand" in tags or 'hand' in tags:
            category = "Hand"
            casecontent = casecontent.replace(' '*4 + '#', '')

        sheetname = _get_ws(export_dir, c[0])

        #print("Get sheete name :"+sheetname)

        # print(suitename,c[1],c[2],casecontent,c[3],category)

        if not sheetname in wb.sheetnames:
            ws = wb.create_sheet(sheetname)
            #ws = wb.active
            ws.append(["Suite_Name", "Case_Name", "Case_Doc",
                       "Case_Content", "Case_Tag", "Case_Type"])
        else:
            ws = wb[sheetname]
            #ws = wb.active
        ws.append([suitename, c[1], c[2], casecontent, c[3], category])

    os.remove(export_file) if os.path.exists(export_file) else None
    wb.save(export_file)
    log.info("生成测试用例文件 {} 到目录 {}".format(export_dir, export_file))

    return (True, export_file)


def export_casexlsy(key, db, exp_filedir=''):
    """
    Download ZHIYAN.com xls Format case from robot case, do zhiyan.com loading.
    :param key: dir
    :param db: case info db
    :param exp_filedir: output dir
    :return: result, file_name
    """

    export_dir = key
    if not os.path.isdir(export_dir):
        log.error("不支持导出一个文件中的用例:" + export_dir)
        return (False, "不支持导出一个文件中的用例:" + export_dir)

    basename = os.path.basename(export_dir)

    dir = exp_filedir
    if dir == '':
        dir = os.environ["AUTO_TEMP"]

    os.mkdir(dir) if not os.path.exists(dir) else None

    export_file = os.path.join(dir, basename + '_rfZHIYAN.xlsx')

    db.refresh_caseinfo(export_dir, "Force")

    cases = []
    sql = "SELECT info_key,info_name,info_doc,info_tags FROM testcase WHERE info_key like '{}%' ;".format(
        key)
    res = db.runsql(sql)
    for i in res:
        (info_key, info_name, info_doc, info_tag) = i
        if info_key.endswith(".robot"):        # Only robot file
            cases.append([info_key, info_name, info_doc, info_tag])

    log.info(f"导出智研用例，目录：{key}, 导出文件:{export_file}")

    wb = Workbook()
    ws = wb.active
    ws.append(["用例标题",
               "所属目录",
               "用例场景",
               "前置条件",
               "操作步骤",
               "预期结果",
               "备注",
               "用例标签",
               "关联TAPD",
               "用例级别",
               "是否模板",
               "执行平台",
               "代码库地址",
               "代码路径",
               "执行命令"])
    for c in cases:
        case_path = c[0]
        case_name = c[1]                                                                           # 用例标题 required
        case_doc  = c[2]
        case_tags = c[3]
        if not os.path.exists(case_path):
            continue
        case_content = getCaseContent(case_path, case_name)
        dir_name = os.path.splitext(case_path)[0].replace(os.environ["PROJECT_DIR"]+'/', '')        # 所属目录 required
        case_scenario = os.path.basename(case_path)
        case_pre_action = "准备测试环境"
        case_exp_result = "执行成功，断言通过"
        case_tapd = os.environ.get("PROJECT_NAME", "")
        case_level = "P0"
        case_is_template = "No"
        case_platform = "tcase"
        case_code_base = "https://git.woa.com/Dollar/DollarPS.git"                                   # 代码库地址 required
        case_code_src = dir_name + ".robot:" + case_name
        case_run_command = "cd $HOME; sh run_zhiyan_robot.sh " + case_path + " --test " + case_name  # 执行命令 required

        ws.append([case_name,
                   dir_name,
                   case_scenario,
                   case_pre_action,
                   case_content,
                   case_exp_result,
                   case_doc,
                   case_tags,
                   case_tapd,
                   case_level,
                   case_is_template,
                   case_platform,
                   case_code_base,
                   case_code_src,
                   case_run_command])

    os.remove(export_file) if os.path.exists(export_file) else None
    wb.save(export_file)
    log.info("生成测试用例文件 {} 到目录 {}".format(export_dir, export_file))
    return True, export_file


def export_casexlsp(key, db, exp_filedir=''):
    """
    下载pytest用例生成符合智研平台的xls文件.可以直接导入 zhiyan.com
    :param key: dir
    :param db: case info db
    :param exp_filedir: output dir
    :return: result, file_name
    """

    export_dir = key
    if not os.path.isdir(export_dir):
        log.error("不支持导出一个文件中的用例:" + export_dir)
        return (False, "不支持导出一个文件中的用例:" + export_dir)

    basename = os.path.basename(export_dir)

    dir = exp_filedir
    if dir == '':
        dir = os.environ["AUTO_TEMP"]

    os.mkdir(dir) if not os.path.exists(dir) else None

    export_file = os.path.join(dir, basename + '_pyZHIYAN.xlsx')

    log.info("开始刷新用例...")
    db.refresh_caseinfo(export_dir, "Force")
    log.info("完成刷新用例...")

    cases = []
    sql = "SELECT info_key,info_name,info_doc,info_tags FROM testcase WHERE info_key like '{}%' ;".format(
        key)
    res = db.runsql(sql)
    for i in res:
        (info_key, info_name, info_doc, info_tag) = i
        if info_key.endswith(".py"):                     # Only pytest file
            cases.append([info_key, info_name, info_doc, info_tag])

    log.info(f"导出智研用例，目录：{key}, 导出文件:{export_file}")

    wb = Workbook()
    ws = wb.active
    ws.append(["用例标题",
               "所属目录",
               "用例场景",
               "前置条件",
               "操作步骤",
               "预期结果",
               "备注",
               "用例标签",
               "关联TAPD",
               "用例级别",
               "是否模板",
               "执行平台",
               "代码库地址",
               "代码路径",
               "执行命令"])
    for c in cases:
        case_path = c[0]
        case_name = c[1]                                                                           # 用例标题 required

        case_tags = c[3]
        if not os.path.exists(case_path):
            continue
        case_content = getCaseDoc_pytest(case_path, case_name)
        case_doc = case_content
        dir_name = os.path.splitext(case_path)[0].replace(os.environ["PROJECT_DIR"]+'/', '')        # 所属目录 required
        case_scenario = os.path.basename(case_path)
        case_pre_action = "准备测试环境"
        case_exp_result = "执行成功，断言通过"
        case_tapd = os.environ.get("PROJECT_NAME", "")
        case_level = "P0"
        case_is_template = "No"
        case_platform = "tcase"
        case_code_base = "https://git.woa.com/Dollar/DollarPS.git"                                   # 代码库地址 required
        case_code_src = dir_name + ".robot:" + case_name
        case_run_command = "cd $HOME; pytest " + case_path + "::" + case_name  # 执行命令 required

        ws.append([case_name,
                   dir_name,
                   case_scenario,
                   case_pre_action,
                   case_content,
                   case_exp_result,
                   case_doc,
                   case_tags,
                   case_tapd,
                   case_level,
                   case_is_template,
                   case_platform,
                   case_code_base,
                   case_code_src,
                   case_run_command])

        #log.info("write case to file: {} : {}".format(case_name, case_path))

    os.remove(export_file) if os.path.exists(export_file) else None
    wb.save(export_file)
    log.info("生成测试用例文件 {} 到目录 {}".format(export_dir, export_file))
    return True, export_file


def _get_ws(export_dir, suite_key):
    """
    return worksheet name
    suite_key= /xxx/project/TestCase/v50/1dir1/test1.robot
    expor_dir= /xxx/project/TestCase
    """

    suite_name = os.path.basename(suite_key)  # test1.robot
    suite_dir = os.path.dirname(suite_key)  # /xxx/project/TestCase/v50/1dir1

    subdir = suite_dir.split(export_dir)[1]  # /v50/1dir1
    subdir = subdir.replace('/', '-')  # _v50_1dir1
    subdir = subdir[1:]  # v50_1dir1

    if subdir == '':
        singal_suite = suite_name.split(".")[0]
        return "_"+singal_suite+"_"

    return subdir


def do_importfromzip(temp_file, path):

    zip_file = temp_file

    try:
        if not os.path.exists(zip_file):
            return ('fail', 'Can not find xlsx file :{}'.format(zip_file))
        if not os.path.isdir(path):
            return ('fail', 'The Node is NOT A DIR :{}'.format(path))

        if not zipfile.is_zipfile(zip_file):
            return ('fail', 'The file is not a zip file :{}'.format(os.path.basename(zip_file)))

        remove_dir(path) if os.path.exists(path) else None
        os.mkdir(path)

        fz = zipfile.ZipFile(zip_file, 'r')
        for file in fz.namelist():
            fz.extract(file, path)

        return ('success', path)
    except Exception as e:
        log.error("从zip文件导入发生异常:{}".format(e))
        return ("fail", "Exception occured .")


def do_unzip_project(temp_file, path):

    zip_file = temp_file

    try:
        if not os.path.exists(zip_file):
            return ('fail', '找不到zip文件:{}'.format(zip_file))

        app = current_app._get_current_object()

        if not zipfile.is_zipfile(zip_file):
            return ('fail', '不是一个zip文件 :{}'.format(os.path.basename(zip_file)))

        remove_dir(path) if os.path.exists(path) else None
        os.mkdir(path)

        fz = zipfile.ZipFile(zip_file, 'r')
        for file in fz.namelist():
            fz.extract(file, path)

        projectfile = ''
        project_content = ''
        for p in os.listdir(path):
            if os.path.exists(os.path.join(path, p, 'platforminterface/project.conf')):
                projectfile = os.path.join(
                    path, p, 'platforminterface/project.conf')
                project_content = os.path.join(path, p)

        if not projectfile:
            msg = "Load Project Fail: 找不到 project.conf:{} ".format(projectfile)
            log.error(msg)
            return ('fail', msg)

        log.info("读取 Project file: {}".format(projectfile))

        with open(projectfile, 'r') as f:
            for l in f:
                if l.startswith('#'):
                    continue
                if len(l.strip()) == 0:
                    continue
                splits = l.strip().split('|')
                if len(splits) != 4:
                    log.error("错误的 project.conf 行 " + l)
                    return ('fail', "错误的 project.conf 行 " + l)
                (projectname, owner, users, cron) = splits
                project_path = os.path.join(
                    app.config['AUTO_HOME'], 'workspace', owner, projectname)
                if os.path.exists(project_path):
                    msg = '目标目录存在:{}'.format(project_path)
                    log.error(msg)
                    return ('fail', msg)
                log.info("复制文件从 {} 到 {} ".format(
                    project_content, project_path))
                try:
                    shutil.copytree(project_content, project_path)
                except Exception as e:
                    return ('fail', "{}".format(e))

        return ('success', project_path)

    except Exception as e:
        log.error("从zip文件导入发生异常:{}".format(e))
        return ("fail", "Exception occured .")


def do_uploadcaserecord(temp_file):

    if not os.path.exists(temp_file):
        return ('fail', 'Can not find file :{}'.format(temp_file))

    app = current_app._get_current_object()

    total = 0
    success = 0
    formaterror = 0
    exits = 0

    with open(temp_file, 'r') as f:
        for l in f:
            l = l.strip()
            if len(l) != 0:
                total += 1
            else:
                continue
            splits = l.split('|')
            if len(splits) != 8:
                formaterror += 1
                log.error("uploadcaserecord 错误到列:"+l)
                continue
            (info_key, info_name, info_testproject, info_projectversion,
             ontime, run_status, run_elapsedtime, run_user) = splits
            sql = ''' INSERT into caserecord (info_key,info_name,info_testproject,info_projectversion,ontime,run_status,run_elapsedtime,run_user)
                      VALUES ('{}','{}','{}','{}','{}','{}','{}','{}');
                      '''.format(info_key, info_name, info_testproject, info_projectversion, ontime, run_status, run_elapsedtime, run_user)
            res = app.config['DB'].runsql(sql)
            if res:
                success += 1
            else:
                exits += 1
                log.error("uploadcaserecord 记录存在:"+l)

    return ('success', 'Finished with total:{}, sucess:{}, error:{}, exists:{}'.format(total, success, formaterror, exits))


def do_importfromxlsx(temp_file, path):

    xls_file = temp_file
    dest_dir = path

    if not os.path.isdir(dest_dir):
        return ('fail', 'The Node is NOT A DIR :{}'.format(dest_dir))
    if not os.path.exists(xls_file):
        return ('fail', 'Can not find xlsx file :{}'.format(xls_file))
    xls_name = os.path.basename(xls_file).split('.')[0]
    dir_name = os.path.basename(dest_dir)
    if not xls_name == dir_name:
        return ('fail', 'Filename {} is not equal to dir name :{}'.format(xls_name, dest_dir))

    try:
        wb = load_workbook(xls_file)

        update_cases = 0
        unupdate_case = 0
        failedlist = []
        for stn in wb.sheetnames[1:]:
            ws = wb[stn]
            if not ws['A1'] != 'Suite_Name':
                return ('fail', 'sheet:{} A1:{} Expect:Suite_Name'.format(stn, ws['A1']))
            if not ws['B1'] != 'Case_Name':
                return ('fail', 'sheet:{} B1:{} Expect:Case_Name'.format(stn, ws['B1']))
            if not ws['C1'] != 'Case_Doc':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Doc'.format(stn, ws['C1']))
            if not ws['D1'] != 'Case_Content':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Content'.format(stn, ws['D1']))
            if not ws['E1'] != 'Case_Tag':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Tag'.format(stn, ws['E1']))
            if not ws['F1'] != 'Case_Type':
                return ('fail', 'sheet:{} C1:{} Expect:Case_Type'.format(stn, ws['F1']))

            for r in ws.rows:
                (a, b, c, d, e, f) = r
                if a.value == 'Suite_Name':    # omit the 1st line
                    continue
                fields = [a.value if a.value else '',
                          b.value if b.value else '',
                          c.value if c.value else '',
                          d.value if d.value else '',
                          e.value if e.value else '',
                          f.value if f.value else ''
                          ]
                (done, msg) = _update_onecase(dest_dir, stn, fields)
                if done:
                    update_cases += 1
                else:
                    unupdate_case += 1
                    failedlist.append(
                        "sheet:{} suite:{} case:{} ->{}".format(stn, a.value, b.value, msg))

        return ('success', 'S:{},F:{},Failist:{}'.format(update_cases, unupdate_case, '\n'.join(failedlist)))

    except Exception as e:
        log.error("do_uploadcase 异常:{}".format(e))
        return ('fail', 'Deal with xlsx file fail :{}'.format(xls_file))


def _update_onecase(dest_dir, sheetname, fields):

    stn = sheetname

    DONE = False

    robotname = fields[0].split('.')[0]
    if stn.startswith('_') and stn.endswith('_'):
        if not '_'+robotname+'_' == stn:
            return (False, "Sheetname Should be same as the First column（no metter ext）:{} vs {}".format(stn, robotname))
        robotfile = os.path.join(dest_dir, robotname+'.robot')
    else:
        subdir = stn.replace('-', '/')
        robotfile = os.path.join(dest_dir, subdir, robotname+'.robot')

    file_dir = os.path.dirname(robotfile)
    os.makedirs(file_dir, exist_ok=True)

    log.info("Updating robotfile:{} with args:{}".format(robotfile, fields))

    isHand = False
    if fields[5] == '手工' or fields[5] == 'HAND' or fields[5] == 'Hand' or fields[5] == 'hand':
        isHand = True

    brandnew = "*** Settings ***\n" + \
               "*** Variables ***\n" + \
               "*** Test Cases ***\n" + \
               "NewTestCase\n" + \
               "    [Documentation]  This is Doc \n" + \
               "    [Tags]   tag1  tag2\n" + \
               "    Log  This is a Brandnew case.\n"

    name = fields[1].strip()
    doc = fields[2].strip()
    content = fields[3].strip()
    tags = fields[4].strip().replace('，', ',').split(',')  # Chinese characters
    if isHand:
        tags.append('Hand')
    tags = list(set(tags))

    space_splitter = re.compile(u'[ \t\xa0]{2,}|\t+')  # robot spliter

    try:
        # 如果文件不存在，直接创建文件和用例
        if not os.path.exists(robotfile):

            log.info("测试用例文件不存在，创建 :"+robotfile)
            with open(robotfile, 'w') as f:
                f.write(brandnew)

            suite = TestSuiteBuilder().build(robotfile)
            t = suite.tests[0]
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n', '\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = "    ".join("", "#*"+l.strip())
                    steps.append(step)
                steps.append("    No Operation")
            else:
                lines = content.split('\n')
                for l in lines:
                    # step = Step(space_splitter.split(l.strip()))
                    step = "    ".join(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps  # 如果用例不存在，则所有内容都更新:New test Case , Update all.

            suite.save(txt_separating_spaces=4)  # TODO suite do not have save function in new RF

            DONE = True

            return (DONE, robotfile)

        # 如果文件存在： 1 用例存在 ，2 用例不存在 :If file exits: 1 case exists ,2 Case doesnt exists.
        suite = TestSuiteBuilder().build(robotfile)

        for t in suite.tests:
            if t.name == name:      # 2用例存在:Case exists
                log.info("用例文件存在，且用例存在，更新: "+name)
                t.tags.value = tags
                t.doc.value = doc.replace('\n', '\\n')

                if isHand:  # 只有手工用例更新 用例内容，存在都自动化用例不更新内容
                    steps = []
                    lines = content.split('\n')
                    for l in lines:
                        step = "    ".join("", "#*" + l.strip())
                        steps.append(step)
                    steps.append("    No Operation")
                    t.steps = steps
                DONE = True
                break

        if DONE:
            suite.save(txt_separating_spaces=4)
            return (DONE, robotfile)

        # 1用例不存在， 需要新增用例: Case doesnt exists， Add new one
        suite = TestSuiteBuilder().build(robotfile)

        if len(suite.tests) > 0:
            log.info("用例文件存在且非空 ,复制并修改成新的用例.")
            t = copy.deepcopy(suite.testcase_table.tests[-1])
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n', '\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = "    ".join("", "#*" + l.strip())
                    steps.append(step)
                steps.append("    No Operation")
            else:
                lines = content.split('\n')
                for l in lines:
                    step = "    ".join(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps

            suite.testcase_table.tests.append(t)

            suite.save(txt_separating_spaces=4)

            DONE = True

            return (DONE, robotfile)

        else:
            # 用例文件存在，但是用例文件里面没有用例，异常
            log.warning("用例文件存在但无内容，删除并新建.")
            os.remove(robotfile)
            with open(robotfile, 'w') as f:
                f.write(brandnew)

            suite = TestSuiteBuilder().build(robotfile)
            t = suite.tests[0]
            t.name = name
            t.tags.value = tags
            t.doc.value = doc.replace('\n', '\\n')

            steps = []
            if isHand:
                lines = content.split('\n')
                for l in lines:
                    step = "    ".join("", "#*"+l.strip())
                    steps.append(step)
                steps.append("    No Operation")
            else:
                lines = content.split('\n')
                for l in lines:
                    step = "    ".join(space_splitter.split(l.strip()))
                    steps.append(step)

            t.steps = steps

            suite.save(txt_separating_spaces=4)

            DONE = True

            return (DONE, robotfile)

    except Exception as e:
        log.error("updateOneCase 异常:{}".format(e))
        return(DONE, "ErrorOccur:{}".format(e))


if __name__ == '__main__':

    #from utils.dbclass import TestDB
    #myDB = TestDB('/Users/tester/PycharmProjects/uniRobotDev/.beats')

    # print(export_cases("/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/tbdsadmin/RobotTbds/TestCase/v41",myDB))
    #print("sss" + _get_ws("/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/Admin/Demo_Project/RobotTestDemo/TestCase/903dir1","/Users/tester/PycharmProjects/uniRobotDev/.beats/workspace/Admin/Demo_Project/RobotTestDemo/TestCase/903dir1/test1.robot")+"xxx")
    print(do_importfromxlsx('/Users/tester/PycharmProjects/uniRobotDev/work/runtime/testloadcase.xlsx',
                            '/Users/tester/PycharmProjects/uniRobotDev/work/workspace/Admin/123TestLoad/testloadcase'))
